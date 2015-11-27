#!/usr/bin/python
# coding=utf-8

from openpyxl import Workbook
from openpyxl import load_workbook

from PyQt4.QtGui import *
from PyQt4.QtCore import *

import requests
from bs4 import BeautifulSoup
import sys
reload(sys)
sys.setdefaultencoding('utf-8')
QTextCodec.setCodecForTr(QTextCodec.codecForName("utf8"))

def update_msg(msg):
    print(msg)

class QueryThread(QThread):
    trigger = pyqtSignal()
    result = "nothing"

    def __int__(self):
        super(QueryThread,self).__init__()

    def render(self,express,TrackingNo):
        self.express = express
        self.TrackingNo = TrackingNo

    def run(self):
        update_msg(u'正在查询...')
        url = 'http://wap.kuaidi100.com/wap_result.jsp?rand=20120517&id=%s&fromWeb=null&&postid=%s'%(self.express,self.TrackingNo)
        try:
            html = requests.get(url,timeout=100).text
            soup = BeautifulSoup(html,'lxml')
            text = str(soup.find_all('p')[3:-1])
            #print text
            if '\u6b64\u5355\u53f7\u6682\u65e0\u7269\u6d41\u4fe1\u606f\uff0c\u8bf7\u7a0d\u540e\u518d\u67e5\u3002' in text:
                self.result = '此单号暂无物流信息，请稍后再查询。'
            elif '\u5355\u53f7\u975e\u6cd5' in text:
                self.result = '单号非法,不足5位或者超出20位。'
            elif '\u5355\u53f7\u4e0d\u6b63\u786e' in text:
                self.result = '单号不正确,单号由12-14位数字字母组成。'
            else:
                for each in soup.find_all('p')[3:-1]:
                    each = str(each)
                    each = each.replace('<p>','')
                    each = each.replace('<strong>','')
                    each = each.replace('</strong>','')
                    each = each.replace('<br/>','')
                    each = each.replace('</p>','')
                    each = each.replace('·','')
                    self.result = each.decode('utf-8')

        except Exception as e:
            QMessageBox.information(None,u'发生异常',str(e))

def openWorkBook(fileName):
    return load_workbook(filename=fileName)

def closeWorkBook(work_book, fileName):
    work_book.save(fileName)


def getWorkSheet(work_book, work_sheet):
    return work_book.get_sheet_by_name(work_sheet)

# 获取物流单号
def getTrackingNo(work_sheet, cell):
    return work_sheet.cell(cell).value

def setTrackingNo(work_sheet, cell, val):
    work_sheet.cell(cell).val = val
def onlyCharNum(s,oth=''):   
    s = s.lower();   
    fomart = 'abcdefghijklmnopqrstuvwxyz013456789'   
    for c in s:   
        if not c in fomart:   
            s = s.replace(c,'');   
    return s; 

def main(fileName):
    queryThread = QueryThread()
    #fileName = 'EmsDataProc.xlsx'

    wb = openWorkBook(fileName)
    ws = getWorkSheet(wb,"Sheet1")

    for r in range(2, ws.max_row):
        TrackingNo = ws.cell(row=r, column=4).value
        queryThread.render('ems',TrackingNo)
        queryThread.run()
        print(queryThread.result)
        ws.cell(row=r, column=10).value = queryThrea.result
        sleep (2)

    closeWorkBook(wb, fileName)


if __name__ == '__main__':
    if len(sys.argv) != 2:
        print 'Usage: python EmsDataProc.py doc'
        exit(1)

    main(sys.argv[1])
